"""
Flask Web 应用
展示: 书籍大纲 + 代码实现说明 + ETH/USDT 多周期回测结果对比 + 策略优化对比
"""

import gzip as _gzip
import glob
import json
import os
import subprocess
import sys
from datetime import timedelta, datetime
from functools import wraps
from io import BytesIO

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from flask import Flask, render_template, jsonify, request, redirect, url_for, session, send_from_directory, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from date_range_report import load_latest_report_from_db
from web_routes import register_page_routes, register_result_api_routes
import config_store

app = Flask(__name__)

# JSON 安全编码: 将 Infinity/NaN 替换为合法 JSON 值
import math as _math

def _sanitize_for_json(obj):
    """递归替换 float inf/nan 为 JSON 安全值"""
    if isinstance(obj, float):
        if _math.isinf(obj):
            return 999.99 if obj > 0 else -999.99
        if _math.isnan(obj):
            return 0.0
        return obj
    if isinstance(obj, dict):
        return {k: _sanitize_for_json(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return type(obj)(_sanitize_for_json(item) for item in obj)
    return obj

class _SafeJSONProvider(app.json_provider_class):
    """在序列化前清理 Infinity/NaN"""
    def dumps(self, obj, **kwargs):
        return super().dumps(_sanitize_for_json(obj), **kwargs)

app.json_provider_class = _SafeJSONProvider
app.json = _SafeJSONProvider(app)

# 生产环境应设置 SECRET_KEY，否则会话可被伪造；未设置时使用临时密钥并告警
_secret = os.environ.get('SECRET_KEY')
app.secret_key = _secret or os.urandom(24).hex()
if not _secret:
    import warnings
    warnings.warn('SECRET_KEY 未设置，使用临时密钥；生产环境请设置 SECRET_KEY', UserWarning)

# ======================================================
#   响应 gzip 压缩中间件（JSON 体积降 ~70%）
# ======================================================
_GZIP_MIN_SIZE = 512  # bytes

@app.after_request
def _gzip_response(response):
    """对 JSON/HTML 响应自动 gzip 压缩。"""
    if (response.status_code < 200 or response.status_code >= 300
            or response.direct_passthrough
            or 'Content-Encoding' in response.headers
            or not response.content_type
            or not response.content_type.startswith(('application/json', 'text/'))):
        return response
    accept = request.headers.get('Accept-Encoding', '')
    if 'gzip' not in accept.lower():
        return response
    data = response.get_data()
    if len(data) < _GZIP_MIN_SIZE:
        return response
    compressed = _gzip.compress(data, compresslevel=6)
    response.set_data(compressed)
    response.headers['Content-Encoding'] = 'gzip'
    response.headers['Content-Length'] = len(compressed)
    response.headers['Vary'] = 'Accept-Encoding'
    return response


# ======================================================
#   用户认证（默认 admin/system123；正式上线时用 ADMIN_PASSWORD 或 ADMIN_PASSWORD_HASH 覆盖）
# ======================================================
def _build_users():
    admin_user = os.environ.get('ADMIN_USER', 'admin')
    pwd = os.environ.get('ADMIN_PASSWORD')
    pwd_hash = os.environ.get('ADMIN_PASSWORD_HASH')
    if pwd_hash:
        return {admin_user: pwd_hash}
    if pwd:
        return {admin_user: generate_password_hash(pwd)}
    return {admin_user: generate_password_hash('system123')}


USERS = _build_users()
SESSION_LIFETIME_DAYS = int(os.environ.get('SESSION_LIFETIME_DAYS', '7'))


def login_required(f):
    """登录验证装饰器"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function


@app.route('/favicon.ico')
def favicon():
    return send_from_directory(os.path.join(app.root_path, 'static'),
                               'favicon.ico', mimetype='image/x-icon')


@app.before_request
def check_auth():
    """全局认证检查 - 除登录页和静态资源外，所有请求都需要认证"""
    # 允许访问登录页面和静态资源
    allowed_endpoints = ('login', 'static', 'favicon')
    if request.endpoint in allowed_endpoints:
        return
    # 未登录
    if not session.get('logged_in'):
        # API 请求返回 JSON 401（而非 HTML 重定向，避免前端 JSON 解析失败）
        if request.path.startswith('/api/'):
            return jsonify({"success": False, "error": "未登录", "login_required": True}), 401
        return redirect(url_for('login', next=request.url))


@app.route('/login', methods=['GET', 'POST'])
def login():
    """登录页面"""
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        if username in USERS and check_password_hash(USERS[username], password):
            session['logged_in'] = True
            session['username'] = username
            session.permanent = True
            app.permanent_session_lifetime = timedelta(days=SESSION_LIFETIME_DAYS)
            next_url = request.args.get('next') or request.form.get('next') or url_for('page_overview')
            return redirect(next_url)
        else:
            error = '用户名或密码错误'
    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    """登出"""
    session.clear()
    return redirect(url_for('login'))

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
BACKTEST_FILE = os.path.join(BASE_DIR, 'backtest_result.json')
BACKTEST_MULTI_FILE = os.path.join(BASE_DIR, 'backtest_multi.json')
GLOBAL_STRATEGY_FILE = os.path.join(BASE_DIR, 'global_strategy_result.json')
STRATEGY_COMPARE_FILE = os.path.join(BASE_DIR, 'strategy_compare_result.json')
STRATEGY_OPTIMIZE_FILE = os.path.join(BASE_DIR, 'strategy_optimize_result.json')
STRATEGY_ENHANCED_FILE = os.path.join(BASE_DIR, 'strategy_enhanced_result.json')
STRATEGY_FUTURES_FILE = os.path.join(BASE_DIR, 'strategy_futures_result.json')
STRATEGY_FUTURES_V2_FILE = os.path.join(BASE_DIR, 'strategy_futures_v2_result.json')
STRATEGY_FUTURES_V3_FILE = os.path.join(BASE_DIR, 'strategy_futures_v3_result.json')
STRATEGY_FUTURES_V4_FILE = os.path.join(BASE_DIR, 'strategy_futures_v4_result.json')
STRATEGY_FUTURES_V5_FILE = os.path.join(BASE_DIR, 'strategy_futures_v5_result.json')
STRATEGY_FUTURES_FINAL_FILE = os.path.join(BASE_DIR, 'strategy_futures_final_result.json')
TIMEFRAME_ANALYSIS_FILE = os.path.join(BASE_DIR, 'timeframe_analysis_result.json')
STRATEGY_15M_FILE = os.path.join(BASE_DIR, 'strategy_15m_result.json')
MA_STRATEGY_FILE = os.path.join(BASE_DIR, 'ma_strategy_result.json')
COMBINED_STRATEGY_FILE = os.path.join(BASE_DIR, 'combined_strategy_result.json')
CANDLESTICK_FILE = os.path.join(BASE_DIR, 'candlestick_result.json')
BOLLINGER_FILE = os.path.join(BASE_DIR, 'bollinger_result.json')
VOLUME_PRICE_FILE = os.path.join(BASE_DIR, 'volume_price_result.json')
FIVE_BOOK_FILE = os.path.join(BASE_DIR, 'five_book_fusion_result.json')
OPTIMIZE_SL_TP_FILE = os.path.join(BASE_DIR, 'optimize_sl_tp_result.json')
SIX_BOOK_FILE = os.path.join(BASE_DIR, 'six_book_fusion_result.json')
KDJ_FILE = os.path.join(BASE_DIR, 'kdj_result.json')
OPTIMIZE_SIX_BOOK_FILE = os.path.join(BASE_DIR, 'optimize_six_book_result.json')
BACKTEST_30D_7D_FILE = os.path.join(BASE_DIR, 'backtest_30d_7d_result.json')
MULTI_TF_BACKTEST_30D_7D_FILE = os.path.join(BASE_DIR, 'data', 'backtests', 'backtest_multi_tf_30d_7d_result.json')
MULTI_TF_DATE_RANGE_DB_FILE = os.path.join(BASE_DIR, 'data', 'backtests', 'multi_tf_date_range_reports.db')
NAKED_KLINE_BACKTEST_FILE = os.path.join(BASE_DIR, 'data', 'backtests', 'naked_kline_backtest_result.json')

RESULT_FILE_PATHS = {
    'BACKTEST_FILE': BACKTEST_FILE,
    'BACKTEST_MULTI_FILE': BACKTEST_MULTI_FILE,
    'GLOBAL_STRATEGY_FILE': GLOBAL_STRATEGY_FILE,
    'STRATEGY_COMPARE_FILE': STRATEGY_COMPARE_FILE,
    'STRATEGY_OPTIMIZE_FILE': STRATEGY_OPTIMIZE_FILE,
    'STRATEGY_ENHANCED_FILE': STRATEGY_ENHANCED_FILE,
    'STRATEGY_FUTURES_FILE': STRATEGY_FUTURES_FILE,
    'STRATEGY_FUTURES_V2_FILE': STRATEGY_FUTURES_V2_FILE,
    'STRATEGY_FUTURES_V3_FILE': STRATEGY_FUTURES_V3_FILE,
    'STRATEGY_FUTURES_V4_FILE': STRATEGY_FUTURES_V4_FILE,
    'STRATEGY_FUTURES_V5_FILE': STRATEGY_FUTURES_V5_FILE,
    'STRATEGY_FUTURES_FINAL_FILE': STRATEGY_FUTURES_FINAL_FILE,
    'TIMEFRAME_ANALYSIS_FILE': TIMEFRAME_ANALYSIS_FILE,
    'STRATEGY_15M_FILE': STRATEGY_15M_FILE,
    'MA_STRATEGY_FILE': MA_STRATEGY_FILE,
    'COMBINED_STRATEGY_FILE': COMBINED_STRATEGY_FILE,
    'CANDLESTICK_FILE': CANDLESTICK_FILE,
    'BOLLINGER_FILE': BOLLINGER_FILE,
    'VOLUME_PRICE_FILE': VOLUME_PRICE_FILE,
    'FIVE_BOOK_FILE': FIVE_BOOK_FILE,
    'SIX_BOOK_FILE': SIX_BOOK_FILE,
    'OPTIMIZE_SL_TP_FILE': OPTIMIZE_SL_TP_FILE,
    'OPTIMIZE_SIX_BOOK_FILE': OPTIMIZE_SIX_BOOK_FILE,
    'BACKTEST_30D_7D_FILE': BACKTEST_30D_7D_FILE,
    'MULTI_TF_BACKTEST_30D_7D_FILE': MULTI_TF_BACKTEST_30D_7D_FILE,
    'NAKED_KLINE_BACKTEST_FILE': NAKED_KLINE_BACKTEST_FILE,
}


def load_json(path):
    if os.path.exists(path):
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return None


# ======================================================
#   页面路由与结果 API 路由（模块化注册）
# ======================================================
register_page_routes(app)
register_result_api_routes(app, load_json=load_json, result_paths=RESULT_FILE_PATHS)

# ── 配置存储: 自动迁移旧 JSON 文件到 DB ──
config_store.ensure_migrated()


@app.route('/api/multi_tf_date_range_report')
def api_multi_tf_date_range_report():
    """读取多周期固定区间回测(DB)的最新结果。"""
    data = load_latest_report_from_db(MULTI_TF_DATE_RANGE_DB_FILE)
    if data:
        return jsonify(data)
    return jsonify({
        "error": "未找到区间回测数据，请先运行 python backtest_multi_tf_date_range_db.py",
        "db_path": MULTI_TF_DATE_RANGE_DB_FILE,
    }), 404


# ── 裸K线交易法 · 逐日盈亏 (从 DB 读取) ──
NAKED_KLINE_DB_FILE = os.path.join(BASE_DIR, 'data', 'backtests', 'naked_kline_backtest.db')
LIVE_MONITOR_RULES_FILE = os.path.join(BASE_DIR, 'data', 'live', 'monitor_rules.json')

DEFAULT_LIVE_MONITOR_RULES = {
    "pf_alert_threshold": 1.00,              # 已实现PF低于该值触发预警
    "pf_min_realized_count": 8,              # 至少多少笔已实现交易才判定PF有效
    "current_loss_streak_alert_days": 3,     # 当前连亏天数预警阈值
    "worst_day_net_alert_usdt": -2000.0,     # 最差单日净PnL预警阈值(USDT)
}


@app.route('/api/naked_kline_daily')
def api_naked_kline_daily():
    """从 SQLite DB 加载裸K线策略最新回测的逐日盈亏数据。"""
    from naked_kline_db import load_latest_run
    data = load_latest_run(NAKED_KLINE_DB_FILE)
    if data:
        return jsonify(data)
    return jsonify({
        "error": "未找到裸K线回测数据，请先运行 python backtest_naked_kline.py",
    }), 404


# ── 多周期联合决策 · 逐日盈亏 (从 DB 读取) ──
MULTI_TF_DAILY_DB_FILE = os.path.join(BASE_DIR, 'data', 'backtests', 'multi_tf_daily_backtest.db')


@app.route('/api/multi_tf_daily')
def api_multi_tf_daily():
    """从 SQLite DB 加载回测数据。支持 ?run_id=N, ?lite=1 轻量模式。"""
    import time as _t
    from multi_tf_daily_db import load_latest_run, load_run_by_id
    t0 = _t.time()
    run_id = request.args.get('run_id', type=int)
    lite = request.args.get('lite', default='0') == '1'
    if run_id:
        data = load_run_by_id(run_id, MULTI_TF_DAILY_DB_FILE,
                              include_trades=not lite)
    else:
        data = load_latest_run(MULTI_TF_DAILY_DB_FILE,
                               include_trades=not lite)
    elapsed = (_t.time() - t0) * 1000
    app.logger.info(f"[perf] api_multi_tf_daily run_id={run_id} lite={lite} -> {elapsed:.0f}ms")
    if data:
        return jsonify(data)
    return jsonify({
        "error": "未找到多周期逐日回测数据，请先运行 python backtest_multi_tf_daily.py",
    }), 404


@app.route('/api/multi_tf_daily/trades')
def api_multi_tf_daily_trades():
    """延迟加载：单独返回某 run 的完整交易记录。"""
    import time as _t
    from multi_tf_daily_db import load_trades_by_run
    t0 = _t.time()
    run_id = request.args.get('run_id', type=int)
    if not run_id:
        return jsonify({"error": "需要 run_id 参数"}), 400
    trades = load_trades_by_run(run_id, MULTI_TF_DAILY_DB_FILE)
    elapsed = (_t.time() - t0) * 1000
    app.logger.info(f"[perf] api_multi_tf_daily_trades run_id={run_id} -> {elapsed:.0f}ms, {len(trades)} trades")
    return jsonify(trades)


@app.route('/api/multi_tf_daily/runs')
def api_multi_tf_daily_runs():
    """列出所有回测版本，用于版本选择器和对比。支持 ?limit=N 分页。"""
    import time as _t
    from multi_tf_daily_db import list_runs
    t0 = _t.time()
    limit = request.args.get('limit', default=0, type=int)
    offset = request.args.get('offset', default=0, type=int)
    runs = list_runs(MULTI_TF_DAILY_DB_FILE, limit=limit, offset=offset)
    elapsed = (_t.time() - t0) * 1000
    app.logger.info(f"[perf] api_multi_tf_daily_runs limit={limit} offset={offset} -> {elapsed:.0f}ms, {len(runs)} rows")
    return jsonify(runs)


@app.route('/api/multi_tf_daily/export_trades')
def api_multi_tf_daily_export_trades():
    """导出完整交易记录为 Excel 文件。支持 ?run_id=N，默认最新。"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from multi_tf_daily_db import load_latest_run, load_run_by_id

    run_id = request.args.get('run_id', type=int)
    if run_id:
        data = load_run_by_id(run_id, MULTI_TF_DAILY_DB_FILE)
    else:
        data = load_latest_run(MULTI_TF_DAILY_DB_FILE)
    if not data:
        return jsonify({"error": "未找到回测数据"}), 404

    trades = data.get('trade_details', [])
    run_meta = data.get('run_meta', {})
    summary = data.get('summary', {})
    daily = data.get('daily_records', [])
    actual_run_id = data.get('run_id', run_id or 'latest')
    tag = data.get('version_tag', '')

    wb = Workbook()

    # ── Sheet 1: 交易记录 ──
    ws = wb.active
    ws.title = "交易记录"

    # 样式定义
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="2B3E50", end_color="2B3E50", fill_type="solid")
    green_font = Font(color="228B22", bold=True)
    red_font = Font(color="DC143C", bold=True)
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )
    center_align = Alignment(horizontal='center', vertical='center')

    # 摘要区
    summary_items = [
        ("回测版本", f"Run #{actual_run_id} {tag}"),
        ("回测区间", f"{run_meta.get('start_date', '')} ~ {run_meta.get('end_date', '')}"),
        ("策略收益", f"{summary.get('total_return_pct', 0):.2f}%"),
        ("最大回撤", f"{summary.get('max_drawdown_pct', 0):.2f}%"),
        ("总交易数", f"{len(trades)}"),
        ("胜率", f"{summary.get('win_rate', 0):.1f}%"),
        ("总费用", f"${summary.get('total_fees', 0):,.2f}"),
    ]
    for i, (label, value) in enumerate(summary_items):
        cell_l = ws.cell(row=i + 1, column=1, value=label)
        cell_l.font = Font(bold=True, size=10)
        ws.cell(row=i + 1, column=2, value=value)

    # 交易表头
    headers = ['#', '时间', '操作', '方向', '市场价', '成交价', '数量',
               '名义价值', '保证金', '杠杆', '手续费', '滑点', 'PnL',
               '账户总值', '原因']
    header_row = len(summary_items) + 2
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # 操作名称映射
    action_map = {
        'OPEN_LONG': '开多', 'CLOSE_LONG': '平多',
        'OPEN_SHORT': '开空', 'CLOSE_SHORT': '平空',
        'LIQUIDATED': '强平', 'SPOT_BUY': '现买',
        'SPOT_SELL': '现卖', 'PARTIAL_TP': '部分止盈',
    }
    dir_map = {'long': '多', 'short': '空'}

    # 写入交易数据
    for i, t in enumerate(trades):
        row = header_row + 1 + i
        pnl = t.get('pnl')
        values = [
            i + 1,
            str(t.get('time', '')).replace('T', ' ')[:19],
            action_map.get(t.get('action', ''), t.get('action', '')),
            dir_map.get(t.get('direction', ''), '-'),
            t.get('market_price'),
            t.get('exec_price'),
            t.get('quantity'),
            t.get('notional_value'),
            t.get('margin'),
            t.get('leverage'),
            t.get('fee'),
            t.get('slippage_cost'),
            pnl,
            t.get('after_total'),
            t.get('reason', ''),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx in (5, 6, 7, 8, 9, 11, 12, 13, 14):
                cell.alignment = center_align
            # PnL 着色
            if col_idx == 13 and pnl is not None:
                cell.font = green_font if pnl > 0 else red_font if pnl < 0 else Font()
                cell.number_format = '#,##0.00'
            # 金额格式
            if col_idx in (5, 6, 8, 9, 14) and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
            if col_idx == 7 and isinstance(val, (int, float)):
                cell.number_format = '0.0000'

    # 自动列宽
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row_idx in range(header_row + 1, header_row + 1 + min(len(trades), 50)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), 40))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 3

    # ── Sheet 2: 逐日盈亏 ──
    if daily:
        ws2 = wb.create_sheet("逐日盈亏")
        day_headers = ['日期', '总资产', '持仓ETH', 'ETH价格', '当日PnL',
                       '当日交易数', '累计收益%']
        for col_idx, h in enumerate(day_headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        initial_equity = daily[0].get('total_value', 200000) if daily else 200000
        for i, d in enumerate(daily):
            row = i + 2
            eq = d.get('total_value', 0)
            cum_ret = (eq / initial_equity - 1) * 100 if initial_equity else 0
            day_pnl = d.get('day_pnl', 0)
            vals = [
                d.get('date', ''),
                eq,
                d.get('eth_qty', 0),
                d.get('eth_price', 0),
                day_pnl,
                d.get('day_trades', 0),
                cum_ret,
            ]
            for col_idx, val in enumerate(vals, 1):
                cell = ws2.cell(row=row, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = center_align
                if col_idx == 2:
                    cell.number_format = '#,##0.00'
                if col_idx == 5:
                    cell.number_format = '#,##0.00'
                    if isinstance(val, (int, float)):
                        cell.font = green_font if val > 0 else red_font if val < 0 else Font()
                if col_idx == 7:
                    cell.number_format = '0.00'

        for col_idx in range(1, len(day_headers) + 1):
            ws2.column_dimensions[ws2.cell(row=1, column=col_idx).column_letter].width = 14

    # 生成文件
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"trades_run{actual_run_id}_{tag.replace(' ', '_')}.xlsx" if tag else f"trades_run{actual_run_id}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


# ── Excel 导出辅助函数 ──
def _excel_styles():
    """返回统一的 Excel 样式字典"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    return {
        'header_font': Font(bold=True, color="FFFFFF", size=10),
        'header_fill': PatternFill(start_color="2B3E50", end_color="2B3E50", fill_type="solid"),
        'green_font': Font(color="228B22", bold=True),
        'red_font': Font(color="DC143C", bold=True),
        'bold_font': Font(bold=True, size=10),
        'title_font': Font(bold=True, size=12, color="2B3E50"),
        'thin_border': Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0'),
        ),
        'center': Alignment(horizontal='center', vertical='center'),
    }


def _load_run_data(run_id_param):
    """加载指定 run 的数据，返回 (data, error_response)"""
    from multi_tf_daily_db import load_latest_run, load_run_by_id
    run_id = request.args.get('run_id', type=int) if run_id_param is None else run_id_param
    if run_id:
        data = load_run_by_id(run_id, MULTI_TF_DAILY_DB_FILE)
    else:
        data = load_latest_run(MULTI_TF_DAILY_DB_FILE)
    if not data:
        return None, (jsonify({"error": "未找到回测数据"}), 404)
    return data, None


def _build_monthly_summary(daily):
    """从逐日记录构建月度汇总数据"""
    month_map = {}
    for d in daily:
        mon = d.get('date', '')[:7]
        if not mon:
            continue
        if mon not in month_map:
            month_map[mon] = {'start': None, 'end': None, 'trades': 0, 'pnl': 0, 'days': 0,
                              'max_dd': 0, 'positive_days': 0}
        entry = month_map[mon]
        if entry['start'] is None:
            entry['start'] = d.get('total_value', 0)
        entry['end'] = d.get('total_value', 0)
        entry['trades'] += d.get('day_trades', 0)
        entry['pnl'] += d.get('day_pnl', 0)
        entry['days'] += 1
        dd = abs(d.get('drawdown_pct', 0))
        if dd > entry['max_dd']:
            entry['max_dd'] = dd
        if (d.get('day_pnl', 0) or 0) > 0:
            entry['positive_days'] += 1
    return month_map


def _write_monthly_sheet(ws, monthly_map, styles):
    """将月度汇总写入一个 worksheet"""
    headers = ['月份', '期初资金', '期末资金', '月收益%', '交易次数',
               '已实现PnL', '交易天数', '盈利天数', '月最大回撤%']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['center']
        cell.border = styles['thin_border']

    months = sorted(monthly_map.keys())
    for i, mon in enumerate(months):
        m = monthly_map[mon]
        ret_pct = ((m['end'] - m['start']) / m['start'] * 100) if m['start'] else 0
        vals = [mon, m['start'], m['end'], ret_pct, m['trades'],
                m['pnl'], m['days'], m['positive_days'], -m['max_dd'] if m['max_dd'] else 0]
        row = i + 2
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = styles['thin_border']
            cell.alignment = styles['center']
            if col in (2, 3) and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
            if col == 4 and isinstance(val, (int, float)):
                cell.number_format = '0.00'
                cell.font = styles['green_font'] if val > 0 else styles['red_font'] if val < 0 else Font()
            if col == 6 and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
                cell.font = styles['green_font'] if val > 0 else styles['red_font'] if val < 0 else Font()
            if col == 9 and isinstance(val, (int, float)):
                cell.number_format = '0.00'
                if val < 0:
                    cell.font = styles['red_font']

    # 合计行
    total_row = len(months) + 2
    ws.cell(row=total_row, column=1, value='合计').font = styles['bold_font']
    total_start = monthly_map[months[0]]['start'] if months else 0
    total_end = monthly_map[months[-1]]['end'] if months else 0
    total_ret = ((total_end - total_start) / total_start * 100) if total_start else 0
    total_trades = sum(m['trades'] for m in monthly_map.values())
    total_pnl = sum(m['pnl'] for m in monthly_map.values())
    total_days = sum(m['days'] for m in monthly_map.values())
    total_pos_days = sum(m['positive_days'] for m in monthly_map.values())
    totals = [None, total_start, total_end, total_ret, total_trades, total_pnl, total_days, total_pos_days, None]
    for col, val in enumerate(totals, 1):
        if val is not None:
            cell = ws.cell(row=total_row, column=col, value=val)
            cell.border = styles['thin_border']
            cell.alignment = styles['center']
            cell.font = styles['bold_font']
            if col in (2, 3, 6) and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
            if col == 4:
                cell.number_format = '0.00'

    # 自动列宽
    widths = [10, 14, 14, 12, 10, 14, 10, 10, 12]
    from openpyxl.utils import get_column_letter
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    return total_row


def _write_summary_sheet(ws, data, styles):
    """将回测汇总信息写入一个 worksheet"""
    from openpyxl.styles import Font
    m = data.get('run_meta', {})
    s = data.get('summary', {})
    snap = data.get('strategy_snapshot', {})
    trades = data.get('trade_details', [])
    tag = data.get('version_tag', '')
    rid = data.get('run_id', '')

    items = [
        ("回测版本", f"Run #{rid} {tag}"),
        ("回测区间", f"{m.get('start_date', '')} ~ {m.get('end_date', '')}"),
        ("杠杆", f"{m.get('leverage', 5)}x"),
        ("策略组合", m.get('combo_name', '')),
        ("", ""),
        ("策略收益", f"{s.get('total_return_pct', 0):.2f}%"),
        ("买入持有", f"{s.get('buy_hold_return_pct', 0):.2f}%"),
        ("Alpha", f"{s.get('alpha_pct', 0):.2f}%"),
        ("最大回撤", f"{s.get('max_drawdown_pct', 0):.2f}%"),
        ("期末资金", f"${s.get('final_capital', 0):,.0f}"),
        ("", ""),
        ("总交易数", str(s.get('total_trades', len(trades)))),
        ("平仓数", str(s.get('close_trades', 0))),
        ("胜率", f"{s.get('win_rate_pct', 0):.1f}%"),
        ("盈亏比", f"{s.get('profit_factor', 0):.2f}"),
        ("总费用", f"${s.get('total_costs', 0):,.2f}"),
        ("费用/本金比", f"{s.get('fee_drag_pct', 0):.2f}%"),
        ("强平次数", str(s.get('liquidations', 0))),
        ("", ""),
        ("趋势保护v3", "ON" if snap.get('use_trend_enhance') else "OFF"),
        ("微结构", "ON" if snap.get('use_microstructure') else "OFF"),
        ("双引擎", "ON" if snap.get('use_dual_engine') else "OFF"),
        ("波动目标", "ON" if snap.get('use_vol_target') else "OFF"),
        ("LiveGate", "ON" if snap.get('use_live_gate') else "OFF"),
        ("Regime", "ON" if snap.get('use_regime_aware') else "OFF"),
        ("风控保护", "ON" if snap.get('use_protections') else "OFF"),
    ]
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 35
    for i, (label, value) in enumerate(items):
        if not label:
            continue
        cell_l = ws.cell(row=i + 1, column=1, value=label)
        cell_l.font = styles['bold_font']
        cell_v = ws.cell(row=i + 1, column=2, value=value)
        if 'ON' == value:
            cell_v.font = Font(color="228B22", bold=True)
        elif 'OFF' == value:
            cell_v.font = Font(color="999999")


def _write_daily_sheet(ws, daily, styles):
    """将逐日盈亏数据写入一个 worksheet"""
    from openpyxl.styles import Font
    headers = ['日期', '账户总值', 'USDT余额', 'ETH数量', 'ETH价值', 'ETH价格',
               '冻结保证金', '多头持仓', '空头持仓', '多头入场价', '空头入场价',
               '多头浮盈', '空头浮盈', '当日PnL', '当日交易数', '累计收益%', '回撤%']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['center']
        cell.border = styles['thin_border']

    initial = daily[0].get('total_value', 200000) if daily else 200000
    for i, d in enumerate(daily):
        row = i + 2
        eq = d.get('total_value', 0)
        cum_ret = (eq / initial - 1) * 100 if initial else 0
        vals = [
            d.get('date', ''), eq,
            d.get('usdt', 0), d.get('eth_qty', 0), d.get('spot_eth_value', 0),
            d.get('eth_price', 0), d.get('frozen_margin', 0),
            1 if d.get('has_long') else 0, 1 if d.get('has_short') else 0,
            d.get('long_entry', ''), d.get('short_entry', ''),
            d.get('long_pnl', 0), d.get('short_pnl', 0),
            d.get('day_pnl', 0), d.get('day_trades', 0),
            cum_ret, -(abs(d.get('drawdown_pct', 0))),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = styles['thin_border']
            cell.alignment = styles['center']
            if col in (2, 3, 5, 7) and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
            if col == 4 and isinstance(val, (int, float)):
                cell.number_format = '0.0000'
            if col == 6 and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
            if col in (10, 11) and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
            if col in (12, 13, 14) and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
                cell.font = styles['green_font'] if val > 0 else styles['red_font'] if val < 0 else Font()
            if col in (16, 17) and isinstance(val, (int, float)):
                cell.number_format = '0.00'
                if col == 17 and val < 0:
                    cell.font = styles['red_font']

    from openpyxl.utils import get_column_letter
    widths = [12, 14, 14, 10, 14, 12, 12, 8, 8, 12, 12, 12, 12, 14, 10, 10, 10]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def _write_trades_sheet(ws, trades, styles):
    """将完整交易记录写入一个 worksheet"""
    from openpyxl.styles import Font
    headers = ['#', '时间', '操作', '方向', '市场价', '成交价', '数量',
               '名义价值', '保证金', '杠杆', '手续费', '滑点', 'PnL',
               '账户总值', '原因']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['center']
        cell.border = styles['thin_border']

    action_map = {
        'OPEN_LONG': '开多', 'CLOSE_LONG': '平多',
        'OPEN_SHORT': '开空', 'CLOSE_SHORT': '平空',
        'LIQUIDATED': '强平', 'SPOT_BUY': '现买',
        'SPOT_SELL': '现卖', 'PARTIAL_TP': '部分止盈',
    }
    dir_map = {'long': '多', 'short': '空'}

    for i, t in enumerate(trades):
        row = i + 2
        pnl = t.get('pnl')
        vals = [
            i + 1,
            str(t.get('time', '')).replace('T', ' ')[:19],
            action_map.get(t.get('action', ''), t.get('action', '')),
            dir_map.get(t.get('direction', ''), '-'),
            t.get('market_price'), t.get('exec_price'), t.get('quantity'),
            t.get('notional_value'), t.get('margin'), t.get('leverage'),
            t.get('fee'), t.get('slippage_cost'), pnl,
            t.get('after_total'), t.get('reason', ''),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = styles['thin_border']
            if col in (5, 6, 7, 8, 9, 11, 12, 13, 14):
                cell.alignment = styles['center']
            if col == 13 and pnl is not None:
                cell.font = styles['green_font'] if pnl > 0 else styles['red_font'] if pnl < 0 else Font()
                cell.number_format = '#,##0.00'
            if col in (5, 6, 8, 9, 14) and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
            if col == 7 and isinstance(val, (int, float)):
                cell.number_format = '0.0000'

    from openpyxl.utils import get_column_letter
    for col in range(1, len(headers) + 1):
        max_len = len(str(headers[col - 1]))
        for row_idx in range(2, min(len(trades) + 2, 52)):
            val = ws.cell(row=row_idx, column=col).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), 40))
        ws.column_dimensions[get_column_letter(col)].width = max_len + 3


@app.route('/api/multi_tf_daily/export_monthly')
def api_multi_tf_daily_export_monthly():
    """导出月度汇总为 Excel 文件。"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font

    data, err = _load_run_data(None)
    if err:
        return err

    daily = data.get('daily_records', [])
    actual_run_id = data.get('run_id', 'latest')
    tag = data.get('version_tag', '')
    styles = _excel_styles()

    wb = Workbook()
    ws = wb.active
    ws.title = "月度汇总"
    _write_monthly_sheet(ws, _build_monthly_summary(daily), styles)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"monthly_run{actual_run_id}_{tag.replace(' ', '_')}.xlsx" if tag else f"monthly_run{actual_run_id}.xlsx"
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@app.route('/api/multi_tf_daily/export_all')
def api_multi_tf_daily_export_all():
    """导出整个页面所有数据为多 Sheet Excel 文件: 回测汇总 + 逐日盈亏 + 月度汇总 + 完整交易记录。"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font

    data, err = _load_run_data(None)
    if err:
        return err

    daily = data.get('daily_records', [])
    trades = data.get('trade_details', [])
    actual_run_id = data.get('run_id', 'latest')
    tag = data.get('version_tag', '')
    styles = _excel_styles()

    wb = Workbook()

    # Sheet 1: 回测汇总
    ws1 = wb.active
    ws1.title = "回测汇总"
    _write_summary_sheet(ws1, data, styles)

    # Sheet 2: 逐日盈亏 (全量字段)
    if daily:
        ws2 = wb.create_sheet("逐日盈亏")
        _write_daily_sheet(ws2, daily, styles)

    # Sheet 3: 月度汇总
    if daily:
        ws3 = wb.create_sheet("月度汇总")
        _write_monthly_sheet(ws3, _build_monthly_summary(daily), styles)

    # Sheet 4: 完整交易记录
    if trades:
        ws4 = wb.create_sheet("交易记录")
        _write_trades_sheet(ws4, trades, styles)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"全量报告_run{actual_run_id}_{tag.replace(' ', '_')}.xlsx" if tag else f"全量报告_run{actual_run_id}.xlsx"
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


# ======================================================
#   当前最新策略说明（菜单最上方）
# ======================================================
@app.route('/strategy/current')
def page_current_strategy():
    return render_template('page_current_strategy.html', active_page='current-strategy')


# ======================================================
#   策略技术文档（Markdown 渲染）
# ======================================================
@app.route('/strategy/tech-doc')
def page_strategy_tech_doc():
    import markdown
    doc_path = os.path.join(os.path.dirname(__file__), 'docs', 'strategy_tech_doc.md')
    try:
        with open(doc_path, 'r', encoding='utf-8') as f:
            md_text = f.read()
    except FileNotFoundError:
        md_text = '> 文档文件不存在，请检查 `docs/strategy_tech_doc.md`'
    html_content = markdown.markdown(
        md_text,
        extensions=['tables', 'fenced_code', 'codehilite', 'toc'],
        extension_configs={'codehilite': {'css_class': 'highlight', 'guess_lang': False}},
    )
    return render_template(
        'page_strategy_tech_doc.html',
        active_page='strategy-tech-doc',
        content=html_content,
        version='10.2',
        last_updated='2026-02-15',
        win_rate='64.9%',
        cpf='2.51',
    )


# ======================================================
#   策略完整规格书（STRATEGY_SPEC.md Markdown 渲染）
# ======================================================
@app.route('/strategy/spec')
def page_strategy_spec():
    import markdown
    spec_path = os.path.join(os.path.dirname(__file__), 'STRATEGY_SPEC.md')
    try:
        with open(spec_path, 'r', encoding='utf-8') as f:
            md_text = f.read()
    except FileNotFoundError:
        md_text = '> 文档文件不存在，请检查 `STRATEGY_SPEC.md`'
    html_content = markdown.markdown(
        md_text,
        extensions=['tables', 'fenced_code', 'codehilite', 'toc'],
        extension_configs={'codehilite': {'css_class': 'highlight', 'guess_lang': False}},
    )
    return render_template(
        'page_strategy_spec.html',
        active_page='strategy-spec',
        content=html_content,
        last_updated='2026-02-15',
    )


# ======================================================
#   策略技术规格书（Codex 版）
# ======================================================
@app.route('/strategy/spec-codex')
def page_strategy_spec_codex():
    import markdown
    spec_path = os.path.join(os.path.dirname(__file__), 'docs', 'strategy_spec_codex.md')
    try:
        with open(spec_path, 'r', encoding='utf-8') as f:
            md_text = f.read()
    except FileNotFoundError:
        md_text = '> 文档文件不存在，请检查 `docs/strategy_spec_codex.md`'
    html_content = markdown.markdown(
        md_text,
        extensions=['tables', 'fenced_code', 'codehilite', 'toc'],
        extension_configs={'codehilite': {'css_class': 'highlight', 'guess_lang': False}},
    )
    return render_template(
        'page_strategy_spec_codex.html',
        active_page='strategy-spec-codex',
        content=html_content,
        last_updated='2026-02-15',
    )


#   实盘控制面板
# ======================================================
@app.route('/strategy/live-control')
def page_live_control():
    return render_template('page_live_control.html', active_page='live-control')


# 存储后台引擎进程信息
_engine_process = {"proc": None, "phase": None, "started_at": None}


@app.route('/api/live/generate_config', methods=['POST'])
def api_live_generate_config():
    """生成配置模板 → 写入 DB"""
    try:
        # 先生成模板到临时文件，再导入 DB
        import tempfile
        tmp_path = os.path.join(tempfile.gettempdir(), 'live_trading_config_tpl.json')
        r = subprocess.run(
            [sys.executable, 'live_runner.py', '--generate-config', '-o', tmp_path],
            capture_output=True, text=True, timeout=15, cwd=BASE_DIR
        )
        config_data = {}
        if os.path.exists(tmp_path):
            with open(tmp_path, 'r') as f:
                config_data = json.load(f)
            # 写入 DB
            config_store.set_live_trading_config(config_data)
            os.remove(tmp_path)
        return jsonify({
            "success": True,
            "message": "配置模板已生成并保存到数据库",
            "output": r.stdout + r.stderr,
            "config": config_data,
        })
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/live/save_config', methods=['POST'])
def api_live_save_config():
    """保存编辑后的配置 → DB"""
    try:
        config_data = request.get_json()
        config_store.set_live_trading_config(config_data)
        # 同时写一份 JSON 文件供引擎进程读取 (过渡期兼容)
        config_path = os.path.join(BASE_DIR, 'live_trading_config.json')
        with open(config_path, 'w') as f:
            json.dump(config_data, f, indent=2, ensure_ascii=False)
        return jsonify({"success": True, "message": "配置已保存"})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/live/load_config')
def api_live_load_config():
    """从 DB 加载当前配置"""
    config_data = config_store.get_live_trading_config_full()
    if config_data:
        return jsonify({"success": True, "config": config_data})
    # Fallback: 尝试旧文件
    config_path = os.path.join(BASE_DIR, 'live_trading_config.json')
    if os.path.exists(config_path):
        with open(config_path, 'r') as f:
            data = json.load(f)
        # 顺便迁移到 DB
        config_store.set_live_trading_config(data)
        return jsonify({"success": True, "config": data})
    return jsonify({"success": False, "message": "配置不存在，请先生成"})


@app.route('/api/ml/status')
def api_ml_status():
    """检查 ML 模型状态"""
    try:
        model_dir = os.path.join(BASE_DIR, 'data', 'ml_models')
        status = {
            'regime_model': False,
            'quantile_model': False,
            'regime_trained_at': None,
            'quantile_trained_at': None,
        }
        vol_path = os.path.join(model_dir, 'vol_regime_model.txt')
        if os.path.exists(vol_path):
            status['regime_model'] = True
            status['regime_trained_at'] = datetime.fromtimestamp(
                os.path.getmtime(vol_path)
            ).strftime('%Y-%m-%d %H:%M')
        q_path = os.path.join(model_dir, 'quantile_config.json')
        if os.path.exists(q_path):
            status['quantile_model'] = True
            status['quantile_trained_at'] = datetime.fromtimestamp(
                os.path.getmtime(q_path)
            ).strftime('%Y-%m-%d %H:%M')
        status['any_model'] = status['regime_model'] or status['quantile_model']
        return jsonify({"success": True, "status": status})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/live/test_signal', methods=['POST'])
def api_live_test_signal():
    """测试信号计算"""
    tf = request.json.get('timeframe', '1h') if request.is_json else '1h'
    try:
        r = subprocess.run(
            [sys.executable, 'live_runner.py', '--test-signal', '--timeframe', tf],
            capture_output=True, text=True, timeout=120, cwd=BASE_DIR
        )
        return jsonify({
            "success": r.returncode == 0,
            "output": r.stdout,
            "error": r.stderr if r.returncode != 0 else "",
        })
    except subprocess.TimeoutExpired:
        return jsonify({"success": False, "output": "", "error": "超时 (120s)"}), 504
    except Exception as e:
        return jsonify({"success": False, "output": "", "error": str(e)}), 500


@app.route('/api/live/test_signal_multi', methods=['POST'])
def api_live_test_signal_multi():
    """多时间框架并行信号检测"""
    data = request.json or {}
    timeframes = data.get('timeframes', ['15m', '1h', '4h', '24h'])  # 与当前策略 decision_timeframes 一致

    # 校验时间框架
    valid_tfs = {'1m','3m','5m','10m','15m','30m','1h','2h','3h','4h',
                 '6h','8h','12h','16h','24h','1d'}
    timeframes = [tf for tf in timeframes if tf in valid_tfs]
    if not timeframes:
        return jsonify({"success": False, "error": "无有效的时间框架"}), 400

    tf_str = ','.join(timeframes)
    output_file = os.path.join(BASE_DIR, 'multi_signal_result.json')

    try:
        # 同步 subprocess 会占用当前 worker 至多 timeout；sync worker 下会阻塞其他请求
        # 超时 = 300s (线上服务器配置较低, 9个周期可能需要2-4分钟)
        r = subprocess.run(
            [sys.executable, 'live_runner.py', '--test-signal-multi',
             '--timeframe', tf_str, '-o', output_file],
            capture_output=True, text=True,
            timeout=300, cwd=BASE_DIR
        )

        # 尝试读取结构化 JSON 结果
        result_data = None
        if os.path.exists(output_file):
            try:
                with open(output_file, 'r') as f:
                    result_data = json.load(f)
            except Exception:
                pass

        # 写入带时间戳的服务器端缓存 (供页面加载时恢复)
        if result_data:
            try:
                from datetime import datetime as _dt
                cache = {
                    "data": result_data,
                    "timeframes": timeframes,
                    "ts": _dt.now().isoformat(),
                    "ts_epoch": __import__('time').time(),
                }
                cache_path = os.path.join(BASE_DIR, 'multi_signal_cache.json')
                with open(cache_path, 'w', encoding='utf-8') as f:
                    json.dump(cache, f, ensure_ascii=False, indent=2)
            except Exception:
                pass

        return jsonify({
            "success": r.returncode == 0,
            "output": r.stdout,
            "error": r.stderr if r.returncode != 0 else "",
            "data": result_data,
        })
    except subprocess.TimeoutExpired:
        return jsonify({"success": False, "output": "", "error": "超时 (300s)"}), 504
    except Exception as e:
        return jsonify({"success": False, "output": "", "error": str(e)}), 500


@app.route('/api/live/signal_cache')
def api_live_signal_cache():
    """读取服务器端缓存的最新多周期检测结果"""
    cache_path = os.path.join(BASE_DIR, 'multi_signal_cache.json')
    if not os.path.exists(cache_path):
        return jsonify({"success": False, "message": "暂无缓存"})
    try:
        with open(cache_path, 'r', encoding='utf-8') as f:
            cache = json.load(f)
        # 计算缓存年龄
        import time as _time
        age_sec = _time.time() - cache.get("ts_epoch", 0)
        cache["age_sec"] = round(age_sec)
        cache["age_text"] = (
            f"{int(age_sec)}秒前" if age_sec < 60
            else f"{int(age_sec/60)}分钟前" if age_sec < 3600
            else f"{int(age_sec/3600)}小时前"
        )
        return jsonify({"success": True, **cache})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})


@app.route('/api/live/test_connection', methods=['POST'])
def api_live_test_connection():
    """测试 API 连接"""
    try:
        r = subprocess.run(
            [sys.executable, 'live_runner.py', '--test-connection'],
            capture_output=True, text=True, timeout=30, cwd=BASE_DIR
        )
        return jsonify({
            "success": r.returncode == 0,
            "output": r.stdout,
            "error": r.stderr if r.returncode != 0 else "",
        })
    except Exception as e:
        return jsonify({"success": False, "output": "", "error": str(e)}), 500


def _detect_engine_process():
    """通过 PID 文件 + 系统进程检测引擎是否在运行。
    优先读取 engine.pid 文件锁（比 pgrep 更可靠），
    如果 PID 文件对应进程存活则返回该进程信息。
    """
    pid_file = os.path.join(BASE_DIR, 'data', 'live', 'engine.pid')

    # 方案 1: 读取 PID 文件 + 验证进程存活
    if os.path.exists(pid_file):
        try:
            import fcntl
            f = open(pid_file, 'r')
            try:
                fcntl.flock(f, fcntl.LOCK_EX | fcntl.LOCK_NB)
                # 能拿到锁 → 说明没有引擎持有它，PID 文件是残留
                fcntl.flock(f, fcntl.LOCK_UN)
                f.close()
            except (IOError, OSError):
                # 拿不到锁 → 有进程正在持有 → 引擎在运行
                f.seek(0)
                pid_str = f.read().strip()
                f.close()
                if pid_str.isdigit():
                    pid = int(pid_str)
                    # 读取 engine_state.json 获取 phase
                    phase = 'paper'
                    try:
                        state_file = os.path.join(BASE_DIR, 'data', 'live', 'engine_state.json')
                        with open(state_file, 'r') as sf:
                            state = json.load(sf)
                        phase = state.get('phase', 'paper')
                    except Exception:
                        pass
                    return {"running": True, "pid": pid, "phase": phase}
        except Exception:
            pass

    # 方案 2: 回退到 pgrep（兼容没有 PID 文件的旧版本）
    try:
        r = subprocess.run(
            ['pgrep', '-af', 'live_runner.py'],
            capture_output=True, text=True, timeout=5
        )
        for line in r.stdout.strip().split('\n'):
            if not line.strip():
                continue
            parts = line.split(None, 1)
            if len(parts) >= 2 and 'live_runner.py' in parts[1]:
                pid = int(parts[0])
                cmd = parts[1]
                # 排除 --test-signal / --status 等非引擎命令
                if any(x in cmd for x in ['--test-signal', '--status',
                                           '--kill-switch', '--generate-config',
                                           '--test-connection']):
                    continue
                # 解析 phase
                phase = 'paper'
                if '--phase' in cmd:
                    tokens = cmd.split()
                    try:
                        idx = tokens.index('--phase')
                        if idx + 1 < len(tokens):
                            phase = tokens[idx + 1]
                    except ValueError:
                        pass
                return {"running": True, "pid": pid, "phase": phase, "cmd": cmd}
    except Exception:
        pass
    return {"running": False}


def _read_latest_balance_from_log():
    """从最新的日志文件读取最后一条 BALANCE 记录"""
    log_dir = os.path.join(BASE_DIR, 'logs', 'live')
    # 尝试 JSONL 文件 (更结构化)
    jsonl_files = sorted(glob.glob(os.path.join(log_dir, 'trade_*.jsonl')), reverse=True)
    if jsonl_files:
        try:
            with open(jsonl_files[0], 'r', encoding='utf-8') as f:
                lines = f.readlines()
            # 从后往前找 BALANCE 记录
            for line in reversed(lines):
                line = line.strip()
                if not line:
                    continue
                try:
                    entry = json.loads(line)
                    if entry.get('level') == 'BALANCE':
                        return entry.get('data', {})
                except json.JSONDecodeError:
                    continue
        except Exception:
            pass
    return None


def _safe_float(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _normalize_live_monitor_rules(raw):
    raw = raw or {}
    rules = dict(DEFAULT_LIVE_MONITOR_RULES)
    # 阈值范围做基础裁剪，避免前端误填导致监控建议失真
    rules["pf_alert_threshold"] = float(max(0.1, min(5.0, _safe_float(raw.get("pf_alert_threshold"), rules["pf_alert_threshold"]))))
    rules["pf_min_realized_count"] = int(max(1, min(200, _safe_float(raw.get("pf_min_realized_count"), rules["pf_min_realized_count"]))))
    rules["current_loss_streak_alert_days"] = int(max(1, min(30, _safe_float(raw.get("current_loss_streak_alert_days"), rules["current_loss_streak_alert_days"]))))
    rules["worst_day_net_alert_usdt"] = float(min(-1.0, _safe_float(raw.get("worst_day_net_alert_usdt"), rules["worst_day_net_alert_usdt"])))
    return rules


def _load_live_monitor_rules():
    # 优先从 DB 读取
    db_rules = config_store.get_monitor_rules()
    if db_rules:
        return _normalize_live_monitor_rules(db_rules)
    # Fallback: 旧 JSON 文件
    try:
        if os.path.exists(LIVE_MONITOR_RULES_FILE):
            with open(LIVE_MONITOR_RULES_FILE, 'r', encoding='utf-8') as f:
                rules = _normalize_live_monitor_rules(json.load(f))
            # 顺便迁移到 DB
            config_store.set_monitor_rules(rules)
            return rules
    except Exception:
        pass
    return dict(DEFAULT_LIVE_MONITOR_RULES)


def _save_live_monitor_rules(rules):
    normalized = _normalize_live_monitor_rules(rules)
    # 写入 DB
    config_store.set_monitor_rules(normalized)
    return normalized


def _parse_log_timestamp(ts_text):
    if not ts_text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S,%f", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(ts_text, fmt)
        except ValueError:
            continue
    return None


def _iter_live_jsonl_entries(days=30):
    """遍历最近 N 天的交易 JSONL 记录（按时间正序）。"""
    log_dir = os.path.join(BASE_DIR, 'logs', 'live')
    files = sorted(glob.glob(os.path.join(log_dir, 'trade_*.jsonl')))
    if not files:
        return []

    if days and days > 0:
        cutoff = datetime.now().date() - timedelta(days=max(1, days) - 1)
        selected = []
        for path in files:
            base = os.path.basename(path)
            try:
                date_token = base.split('_', 1)[1].split('.', 1)[0]
                f_date = datetime.strptime(date_token, "%Y%m%d").date()
                if f_date >= cutoff:
                    selected.append(path)
            except Exception:
                # 文件名异常时保留，避免漏数据
                selected.append(path)
        files = selected

    entries = []
    for path in files:
        try:
            with open(path, 'r', encoding='utf-8', errors='replace') as f:
                for raw in f:
                    line = raw.strip()
                    if not line:
                        continue
                    try:
                        item = json.loads(line)
                        entries.append(item)
                    except json.JSONDecodeError:
                        continue
        except Exception:
            continue
    return entries


def _build_live_history_payload(days=30, limit=120):
    """聚合实盘日志: 历史交易 / 逐日盈亏 / 交易统计。"""
    entries = _iter_live_jsonl_entries(days=days)

    daily_map = {}
    trade_records = []
    latest_balance = None
    latest_balance_sort = None

    def get_day_row(date_key):
        row = daily_map.get(date_key)
        if row is None:
            row = {
                "date": date_key,
                "trade_count": 0,
                "realized_count": 0,
                "wins": 0,
                "losses": 0,
                "realized_pnl": 0.0,
                "fees": 0.0,
                "slippage": 0.0,
                "funding": 0.0,
                "open_equity": None,
                "close_equity": None,
                "high_equity": None,
                "low_equity": None,
            }
            daily_map[date_key] = row
        return row

    for entry in entries:
        level = str(entry.get("level", "")).upper()
        data = entry.get("data") or {}
        ts_text = entry.get("timestamp", "")
        dt = _parse_log_timestamp(ts_text)
        sort_key = dt.timestamp() if dt else 0.0
        date_key = dt.date().isoformat() if dt else (ts_text[:10] if ts_text else "unknown")
        day = get_day_row(date_key)

        if level == "TRADE":
            action = str(data.get("action", "")).upper()
            pnl = _safe_float(data.get("pnl"), 0.0)
            fee = _safe_float(data.get("fee"), 0.0)
            slippage = _safe_float(data.get("slippage"), 0.0)
            price = _safe_float(data.get("price"), 0.0)
            qty = _safe_float(data.get("qty"), 0.0)
            notional = price * qty

            day["trade_count"] += 1
            day["fees"] += fee
            day["slippage"] += slippage

            is_realized = (
                action.startswith("CLOSE")
                or action.startswith("PARTIAL")
                or action in {"STOP_LOSS", "LIQUIDATION"}
                or abs(pnl) > 1e-12
            )
            if is_realized:
                day["realized_count"] += 1
                day["realized_pnl"] += pnl
                if pnl > 0:
                    day["wins"] += 1
                elif pnl < 0:
                    day["losses"] += 1

            trade_records.append({
                "timestamp": ts_text,
                "date": date_key,
                "action": action or "--",
                "symbol": data.get("symbol", "--"),
                "side": data.get("side", "--"),
                "price": price,
                "qty": qty,
                "notional": notional,
                "pnl": pnl,
                "fee": fee,
                "slippage": slippage,
                "leverage": int(_safe_float(data.get("leverage"), 0)),
                "reason": data.get("reason", ""),
                "order_id": data.get("order_id", ""),
                "_sort_key": sort_key,
            })

        elif level == "BALANCE":
            equity = _safe_float(data.get("total_equity"), None)
            if equity is None:
                continue
            if day["open_equity"] is None:
                day["open_equity"] = equity
            day["close_equity"] = equity
            day["high_equity"] = equity if day["high_equity"] is None else max(day["high_equity"], equity)
            day["low_equity"] = equity if day["low_equity"] is None else min(day["low_equity"], equity)

            if latest_balance_sort is None or sort_key >= latest_balance_sort:
                latest_balance_sort = sort_key
                latest_balance = {
                    "timestamp": ts_text,
                    "total_equity": equity,
                    "usdt": _safe_float(data.get("usdt"), 0.0),
                    "unrealized_pnl": _safe_float(data.get("unrealized_pnl"), 0.0),
                    "frozen_margin": _safe_float(data.get("frozen_margin"), 0.0),
                    "available_margin": _safe_float(data.get("available_margin"), 0.0),
                    "positions": data.get("positions", []),
                }

        elif level == "FUNDING":
            day["funding"] += _safe_float(data.get("amount"), 0.0)

    daily_records = []
    for date_key in sorted(daily_map.keys(), reverse=True):
        row = daily_map[date_key]
        realized_count = row["realized_count"]
        row["win_rate_pct"] = (row["wins"] / realized_count * 100.0) if realized_count > 0 else 0.0
        # 口径说明: realized_pnl 来自日志中的 pnl，fees/funding 单独展示；net_pnl 仅供监控近似使用
        row["net_pnl"] = row["realized_pnl"] - row["fees"] + row["funding"]
        daily_records.append(row)

    trade_records.sort(key=lambda x: x["_sort_key"], reverse=True)
    for r in trade_records:
        r.pop("_sort_key", None)
    recent_trades = trade_records[:max(20, min(limit, 500))]

    realized_pnls = []
    for r in trade_records:
        pnl = _safe_float(r.get("pnl"), 0.0)
        action = str(r.get("action", ""))
        is_realized = (
            action.startswith("CLOSE")
            or action.startswith("PARTIAL")
            or action in {"STOP_LOSS", "LIQUIDATION"}
            or abs(pnl) > 1e-12
        )
        if is_realized:
            realized_pnls.append(pnl)

    gross_profit = sum(p for p in realized_pnls if p > 0)
    gross_loss = abs(sum(p for p in realized_pnls if p < 0))
    realized_pf = (gross_profit / gross_loss) if gross_loss > 1e-12 else None

    total_realized = sum(r["realized_pnl"] for r in daily_records)
    total_fees = sum(r["fees"] for r in daily_records)
    total_slippage = sum(r["slippage"] for r in daily_records)
    total_funding = sum(r["funding"] for r in daily_records)
    total_realized_count = sum(r["realized_count"] for r in daily_records)
    total_wins = sum(r["wins"] for r in daily_records)
    total_losses = sum(r["losses"] for r in daily_records)
    total_trade_count = len(trade_records)
    total_net = total_realized - total_fees + total_funding

    daily_sorted = sorted(daily_records, key=lambda x: x.get("date", ""))
    daily_net_values = [float(r.get("net_pnl", 0.0)) for r in daily_sorted]
    positive_days = sum(1 for v in daily_net_values if v > 0)
    negative_days = sum(1 for v in daily_net_values if v < 0)
    best_day_net = max(daily_net_values) if daily_net_values else 0.0
    worst_day_net = min(daily_net_values) if daily_net_values else 0.0
    avg_daily_net = (sum(daily_net_values) / len(daily_net_values)) if daily_net_values else 0.0

    max_consecutive_loss_days = 0
    current_loss_streak = 0
    for v in daily_net_values:
        if v < 0:
            current_loss_streak += 1
            if current_loss_streak > max_consecutive_loss_days:
                max_consecutive_loss_days = current_loss_streak
        else:
            current_loss_streak = 0

    current_consecutive_loss_days = 0
    for v in reversed(daily_net_values):
        if v < 0:
            current_consecutive_loss_days += 1
        else:
            break

    fee_drag_pct = (total_fees / gross_profit * 100.0) if gross_profit > 1e-12 else 0.0

    stats = {
        "days": days,
        "total_trade_count": total_trade_count,
        "total_realized_count": total_realized_count,
        "wins": total_wins,
        "losses": total_losses,
        "win_rate_pct": (total_wins / total_realized_count * 100.0) if total_realized_count > 0 else 0.0,
        "realized_pnl": total_realized,
        "fees": total_fees,
        "slippage": total_slippage,
        "funding": total_funding,
        "net_pnl": total_net,
        "avg_realized_pnl": (total_realized / total_realized_count) if total_realized_count > 0 else 0.0,
        "gross_profit": gross_profit,
        "gross_loss": gross_loss,
        "realized_pf": realized_pf,
        "fee_drag_pct": fee_drag_pct,
        "avg_daily_net_pnl": avg_daily_net,
        "best_day_net_pnl": best_day_net,
        "worst_day_net_pnl": worst_day_net,
        "positive_days": positive_days,
        "negative_days": negative_days,
        "max_consecutive_loss_days": max_consecutive_loss_days,
        "current_consecutive_loss_days": current_consecutive_loss_days,
    }

    return {
        "trade_stats": stats,
        "daily_records": daily_records,
        "trade_records": recent_trades,
        "latest_balance": latest_balance,
        "monitor_rules": _load_live_monitor_rules(),
    }


@app.route('/api/live/status')
def api_live_status():
    """获取引擎状态 - 通过文件 + 进程检测，不依赖内存变量"""
    data_dir = os.path.join(BASE_DIR, 'data', 'live')
    result = {
        "engine": None,
        "risk": None,
        "performance": None,
        "performance_summary": None,
        "process": None,
    }

    # 1. 引擎状态文件
    engine_file = os.path.join(data_dir, 'engine_state.json')
    if os.path.exists(engine_file):
        try:
            with open(engine_file, 'r') as f:
                result["engine"] = json.load(f)
        except (json.JSONDecodeError, IOError):
            pass

    # 2. 风控状态文件
    risk_file = os.path.join(data_dir, 'risk_state.json')
    if os.path.exists(risk_file):
        try:
            with open(risk_file, 'r') as f:
                result["risk"] = json.load(f)
        except (json.JSONDecodeError, IOError):
            pass

    # 3. 绩效数据文件
    perf_file = os.path.join(data_dir, 'performance.json')
    if os.path.exists(perf_file):
        try:
            with open(perf_file, 'r') as f:
                perf_raw = json.load(f)
            result["performance"] = perf_raw

            # 统一口径: 额外提供 summary 字段，避免前端自行推导时单位不一致
            try:
                from performance_tracker import PerformanceTracker
                tracker = PerformanceTracker(
                    initial_capital=perf_raw.get("initial_capital", 0),
                    data_dir=data_dir,
                )
                summary = tracker.get_summary()
                result["performance_summary"] = summary

                # 兼容旧前端: 将常用汇总字段回填到 performance 顶层
                merged = dict(perf_raw)
                for k in [
                    "initial_capital", "current_equity", "total_return",
                    "total_return_pct", "total_pnl", "total_trades",
                    "wins", "losses", "win_rate", "win_rate_pct",
                    "max_drawdown", "max_drawdown_pct", "total_fees",
                ]:
                    if k in summary:
                        merged[k] = summary[k]
                result["performance"] = merged
            except Exception:
                pass
        except (json.JSONDecodeError, IOError):
            pass

    # 4. 通过进程检测引擎运行状态 (不依赖内存)
    process_info = _detect_engine_process()
    result["process"] = process_info

    # 5. 如果没有 engine_state 文件但引擎在运行，从日志读取余额
    if result["engine"] is None and process_info.get("running"):
        balance = _read_latest_balance_from_log()
        if balance:
            result["engine"] = {
                "usdt": balance.get("usdt", 0),
                "frozen_margin": balance.get("frozen_margin", 0),
                "equity": balance.get("total_equity", 0),
                "unrealized_pnl": balance.get("unrealized_pnl", 0),
                "positions": balance.get("positions", []),
                "phase": process_info.get("phase", "unknown"),
                "source": "log",  # 标记数据来源
            }

    # 6. 也检查内存中的引擎进程 (Web 启动的)
    proc = _engine_process.get("proc")
    if proc is not None:
        poll = proc.poll()
        if poll is None and not process_info.get("running"):
            # 内存中有进程但 pgrep 没检测到 (不太可能，兜底)
            result["process"] = {
                "running": True,
                "pid": proc.pid,
                "phase": _engine_process.get("phase"),
                "started_at": _engine_process.get("started_at"),
            }

    return jsonify({"success": True, **result})


@app.route('/api/live/history')
def api_live_history():
    """实盘历史聚合: 交易记录 + 逐日盈亏 + 统计"""
    days = request.args.get('days', 30, type=int) or 30
    limit = request.args.get('limit', 120, type=int) or 120
    days = max(1, min(days, 180))
    limit = max(20, min(limit, 500))
    payload = _build_live_history_payload(days=days, limit=limit)
    return jsonify({"success": True, **payload})


@app.route('/api/live/reset-stats', methods=['POST'])
def api_live_reset_stats():
    """重置风控/绩效统计 (资金变更或重新开始时调用)"""
    data_dir = os.path.join(BASE_DIR, 'data', 'live')

    # 可选: 传入新初始资金
    body = request.get_json(silent=True) or {}
    new_capital = float(body.get('new_capital', 0))

    reset_items = []

    # 1. 重置 risk_state.json
    risk_file = os.path.join(data_dir, 'risk_state.json')
    if os.path.exists(risk_file):
        os.remove(risk_file)
        reset_items.append('risk_state.json')

    # 2. 重置 performance.json
    perf_file = os.path.join(data_dir, 'performance.json')
    if os.path.exists(perf_file):
        os.remove(perf_file)
        reset_items.append('performance.json')

    # 3. 重置 engine_state.json 中的累计字段 (保留引擎运行状态)
    engine_file = os.path.join(data_dir, 'engine_state.json')
    if os.path.exists(engine_file):
        try:
            with open(engine_file, 'r') as f:
                eng = json.load(f)
            # 清理累计统计, 保留当前余额和运行状态
            if new_capital > 0:
                eng['usdt'] = new_capital
                eng['equity'] = new_capital
                eng['initial_capital'] = new_capital
            eng['total_pnl'] = 0
            eng['realized_pnl'] = 0
            eng['unrealized_pnl'] = 0
            eng['frozen_margin'] = 0
            eng['positions'] = {}
            eng['total_trades'] = 0
            with open(engine_file, 'w') as f:
                json.dump(eng, f, indent=2, ensure_ascii=False)
            reset_items.append('engine_state.json (统计已清零)')
        except (json.JSONDecodeError, IOError):
            pass

    msg = f"已重置: {', '.join(reset_items)}" if reset_items else "无文件需要重置"
    if new_capital > 0:
        msg += f" (新初始资金: ${new_capital:,.2f})"

    return jsonify({"success": True, "message": msg, "reset_items": reset_items})


@app.route('/api/live/monitor_rules')
def api_live_monitor_rules_get():
    """读取实盘监控规则阈值。"""
    return jsonify({"success": True, "rules": _load_live_monitor_rules()})


@app.route('/api/live/monitor_rules', methods=['POST'])
def api_live_monitor_rules_save():
    """保存实盘监控规则阈值。"""
    data = request.get_json(silent=True) or {}
    raw_rules = data.get("rules", data)
    try:
        saved = _save_live_monitor_rules(raw_rules)
        return jsonify({"success": True, "rules": saved, "message": "监控阈值已保存"})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/live/start', methods=['POST'])
def api_live_start():
    """启动交易引擎"""
    data = request.get_json() or {}
    phase = data.get('phase', 'paper')

    # 检查是否已有引擎运行 (通过进程检测，不仅仅看内存)
    existing = _detect_engine_process()
    if existing.get("running"):
        return jsonify({
            "success": False,
            "message": f"引擎已在运行 (PID={existing['pid']}, phase={existing.get('phase', '?')})"
        })
    proc = _engine_process.get("proc")
    if proc is not None and proc.poll() is None:
        return jsonify({
            "success": False,
            "message": f"引擎已在运行 (PID={proc.pid}, phase={_engine_process['phase']})"
        })

    valid_phases = ['paper', 'testnet', 'small_live', 'scale_up']
    if phase not in valid_phases:
        return jsonify({"success": False, "message": f"无效阶段: {phase}"})

    try:
        # 引擎现在优先从 DB 加载配置, 不再需要传 --config 文件
        cmd = [sys.executable, 'live_runner.py', '--phase', phase, '-y']

        log_dir = os.path.join(BASE_DIR, 'logs', 'live')
        os.makedirs(log_dir, exist_ok=True)
        log_file = open(os.path.join(log_dir, f'engine_{phase}.log'), 'a')

        p = subprocess.Popen(
            cmd, cwd=BASE_DIR,
            stdout=log_file, stderr=subprocess.STDOUT,
            start_new_session=True,   # 脱离 gunicorn 进程树，Web 服务重启不影响引擎
        )
        _engine_process["proc"] = p
        _engine_process["phase"] = phase
        _engine_process["started_at"] = __import__('datetime').datetime.now().isoformat()

        return jsonify({
            "success": True,
            "message": f"引擎已启动 (PID={p.pid}, phase={phase})",
            "pid": p.pid,
        })
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


def _kill_all_engine_processes(exclude_pid=None):
    """杀死所有 live_runner.py 引擎进程（排除指定 PID）。
    用于清理因历史 bug 残留的僵尸进程。"""
    killed = []
    try:
        r = subprocess.run(
            ['pgrep', '-af', 'live_runner.py'],
            capture_output=True, text=True, timeout=5
        )
        for line in r.stdout.strip().split('\n'):
            if not line.strip():
                continue
            parts = line.split(None, 1)
            if len(parts) < 2 or 'live_runner.py' not in parts[1]:
                continue
            cmd = parts[1]
            # 排除非引擎命令
            if any(x in cmd for x in ['--test-signal', '--status',
                                       '--kill-switch', '--generate-config',
                                       '--test-connection']):
                continue
            pid = int(parts[0])
            if exclude_pid and pid == exclude_pid:
                continue
            try:
                os.kill(pid, __import__('signal').SIGKILL)
                killed.append(pid)
            except ProcessLookupError:
                pass
    except Exception:
        pass
    return killed


@app.route('/api/live/stop', methods=['POST'])
def api_live_stop():
    """停止交易引擎 — 支持停止任何方式启动的引擎进程"""
    import signal as sig

    # 优先用内存中的 proc 对象
    proc = _engine_process.get("proc")
    if proc is not None and proc.poll() is None:
        try:
            proc.send_signal(sig.SIGTERM)
            proc.wait(timeout=10)
            _engine_process["proc"] = None
            # 同时清理可能的残留进程
            orphans = _kill_all_engine_processes()
            msg = "引擎已停止"
            if orphans:
                msg += f" (同时清理了 {len(orphans)} 个残留进程)"
            return jsonify({"success": True, "message": msg})
        except subprocess.TimeoutExpired:
            proc.kill()
            _engine_process["proc"] = None
            orphans = _kill_all_engine_processes()
            return jsonify({"success": True, "message": "引擎已强制停止"})
        except Exception as e:
            return jsonify({"success": False, "message": str(e)}), 500

    # proc 为空 (Web 服务重启过, 但引擎仍在运行) → 通过 PID 文件/pgrep 检测
    existing = _detect_engine_process()
    if not existing.get("running"):
        # 兜底: 即使检测不到，也尝试清理残留进程
        orphans = _kill_all_engine_processes()
        _engine_process["proc"] = None
        if orphans:
            return jsonify({"success": True,
                            "message": f"清理了 {len(orphans)} 个残留引擎进程: {orphans}"})
        return jsonify({"success": False, "message": "引擎未在运行"})

    pid = existing["pid"]
    try:
        os.kill(pid, sig.SIGTERM)
        # 等待最多10秒
        import time
        for _ in range(20):
            time.sleep(0.5)
            check = _detect_engine_process()
            if not check.get("running"):
                _engine_process["proc"] = None
                # 清理其他可能的残留进程
                orphans = _kill_all_engine_processes()
                msg = f"引擎已停止 (PID={pid})"
                if orphans:
                    msg += f" + 清理 {len(orphans)} 个残留进程"
                return jsonify({"success": True, "message": msg})
        # 超时 → 强制 kill 所有引擎进程
        _kill_all_engine_processes()
        _engine_process["proc"] = None
        return jsonify({"success": True, "message": f"引擎已强制停止 (PID={pid})"})
    except ProcessLookupError:
        _engine_process["proc"] = None
        _kill_all_engine_processes()
        return jsonify({"success": True, "message": f"引擎进程已不存在 (PID={pid})"})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/live/kill_switch', methods=['POST'])
def api_live_kill_switch():
    """紧急平仓"""
    try:
        # 引擎从 DB 加载配置
        cmd = [sys.executable, 'live_runner.py', '--kill-switch', '-y']

        r = subprocess.run(cmd, capture_output=True, text=True, timeout=30, cwd=BASE_DIR)
        return jsonify({
            "success": r.returncode == 0,
            "output": r.stdout,
            "error": r.stderr if r.returncode != 0 else "",
        })
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/live/logs')
def api_live_logs():
    """获取最近日志"""
    log_dir = os.path.join(BASE_DIR, 'logs', 'live')
    lines = request.args.get('lines', 100, type=int)

    # 找到最新的日志文件
    log_files = sorted(glob.glob(os.path.join(log_dir, 'trade_*.log')), reverse=True)
    if not log_files:
        # 尝试引擎日志
        log_files = sorted(glob.glob(os.path.join(log_dir, 'engine_*.log')), reverse=True)

    if not log_files:
        return jsonify({"success": True, "logs": "暂无日志", "file": ""})

    try:
        with open(log_files[0], 'r', encoding='utf-8', errors='replace') as f:
            all_lines = f.readlines()
            recent = all_lines[-lines:]
        return jsonify({
            "success": True,
            "logs": "".join(recent),
            "file": os.path.basename(log_files[0]),
            "total_lines": len(all_lines),
        })
    except Exception as e:
        return jsonify({"success": False, "logs": str(e), "file": ""})


if __name__ == '__main__':
    print("\n" + "=" * 60)
    print("  技术分析 Web 展示平台")
    print("  五书融合: 背离+均线+蜡烛图+布林带+量价")
    print("  访问: http://127.0.0.1:5000")
    print("=" * 60 + "\n")
    # 排除引擎产生的日志/数据文件，防止 reloader 因这些文件变化而反复重启 Flask，
    # 导致 _engine_process 内存引用丢失、Web UI 误判引擎已停止。
    # 线上用 gunicorn 没有 reloader，不受此影响。
    app.run(
        debug=True,
        port=5000,
        exclude_patterns=[
            '**/logs/**',
            '**/data/**',
            '**/*.log',
            '**/*.jsonl',
            '**/*.pyc',
            '**/__pycache__/**',
            '**/backtest_*_result.json',
            '**/optimize_*_result.json',
        ],
    )
